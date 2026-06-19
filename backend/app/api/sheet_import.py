"""Google Sheets → Domain Manager entity importer (CSV-public).

Flow:
1. POST /api/sheet-import/discover { url } → list of tabs (gid + name)
2. POST /api/sheet-import/preview  { url, gid } → headers + first N rows + auto-detected mapping
3. POST /api/sheet-import/run      { url, gid, target, column_map }
   - target = "servers" | "mail" | "notes"
   - upserts by natural key (host/ip for servers, email for mail)
"""
from __future__ import annotations

import re
from typing import Dict, Any, List, Optional

import json
import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import get_current_user
from app.core.crypto import encrypt_secret
from app.db.session import get_db
from app.models.models import User, RemoteServer, MailAccount, Note, Spreadsheet
from app.services.google_sheets import (
    discover_tabs, fetch_tab, extract_sheet_id, fetch_tab_csv, parse_csv,
)


router = APIRouter(prefix="/api/sheet-import", tags=["sheet-import"])


# ── Schemas ─────────────────────────────────────────────────────────────

class DiscoverIn(BaseModel):
    url: str


class PreviewIn(BaseModel):
    url: str
    gid: str
    limit: int = 20


class RunIn(BaseModel):
    url: str
    gid: str
    target: str  # servers | mail | notes
    column_map: Dict[str, str]   # entity_field → sheet header
    tab_name: Optional[str] = None  # used for notes title


class LocalPreviewIn(BaseModel):
    sheet_id: int
    sheet_index: int = 0


class LocalRunIn(BaseModel):
    sheet_id: int
    sheet_index: int = 0
    target: str  # servers | mail | notes
    column_map: Dict[str, str]
    tab_name: Optional[str] = None


class BatchItem(BaseModel):
    gid: str
    target: str                       # servers | mail | notes
    column_map: Dict[str, str]
    tab_name: Optional[str] = None


class BatchRunIn(BaseModel):
    url: str
    items: List[BatchItem]


class BatchDiscoverIn(BaseModel):
    url: str
    preview_limit: int = 5


# ── Heuristic column auto-mapping ───────────────────────────────────────

# entity_field → list of candidate header substrings (lower-case)
SERVER_HINTS = {
    "host":     ["ip", "host", "адреса", "адрес", "ip:port"],
    "label":    ["name", "label", "назва", "название", "сервер", "регіон", "регион"],
    "username": ["login", "user", "користувач", "пользователь", "username"],
    "password": ["pass", "пароль"],
    "web_url":  ["url", "панель", "panel", "billing", "white", "веб"],
    "notes":    ["geo", "country", "примітк", "примечан", "комент"],
    "tags":     ["tag", "тег"],
}
MAIL_HINTS = {
    "email":    ["email", "mail", "пошт", "почт", "e-mail"],
    "password": ["pass", "пароль"],
    "label":    ["name", "label", "назва"],
    "tags":     ["tag", "тег"],
    "notes":    ["примітк", "примечан", "коммент", "коммент", "імап", "imap"],
}


# ── Value-type classifier (looks at the data itself) ────────────────────

EMAIL_VRE  = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")
IPV4_VRE   = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}(?::\d{1,5})?$")
URL_VRE    = re.compile(r"^https?://\S+$", re.I)
DOMAIN_VRE = re.compile(r"^(?=.{4,253}$)[a-z0-9]([a-z0-9-]*[a-z0-9])?(\.[a-z0-9]([a-z0-9-]*[a-z0-9])?)+$", re.I)
PHONE_VRE  = re.compile(r"^\+?[\d\-\s()]{7,}$")


def classify_value(v: str) -> str:
    """Return a coarse type label for a single cell value."""
    if v is None: return "empty"
    s = str(v).strip()
    if not s: return "empty"
    if EMAIL_VRE.match(s): return "email"
    if IPV4_VRE.match(s):  return "ip"
    if URL_VRE.match(s):   return "url"
    if DOMAIN_VRE.match(s) and not s[0].isdigit() and "." in s:
        return "domain"
    # Phone before password — phones contain digits + dashes/spaces too
    if PHONE_VRE.match(s) and sum(c.isdigit() for c in s) >= 7 and any(c in "+- ()" for c in s):
        return "phone"
    # Password heuristic: needs LENGTH ≥ 8 AND high entropy (≥2 character classes)
    # Avoids false-positives on labels like "337_wawada" (alphanumeric only, no special).
    # A label like "337_wawada" has letter + digit but no special; we now require:
    #   - length ≥ 8 AND
    #   - either (letter + digit + special) OR length ≥ 12 with letter + digit
    if 8 <= len(s) <= 64 and " " not in s:
        has_letter = any(c.isalpha() for c in s)
        has_digit  = any(c.isdigit() for c in s)
        has_special = any(not c.isalnum() and c != "_" for c in s)
        # Strong: 3 classes
        if has_letter and has_digit and has_special: return "password"
        # Medium: 2 classes but quite long
        if has_letter and has_digit and len(s) >= 12: return "password"
        # Mostly random — contains base64/hex-like patterns
        if has_special and (has_letter or has_digit) and len(s) >= 10: return "password"
    return "text"


def column_types(headers: List[str], rows: List[Dict[str, str]]) -> Dict[str, str]:
    """For each header, pick the dominant non-empty value type across rows.
    Falls back to 'empty' if column has no data."""
    out: Dict[str, str] = {}
    for h in headers:
        counts: Dict[str, int] = {}
        for r in rows:
            t = classify_value(r.get(h, ""))
            counts[t] = counts.get(t, 0) + 1
        non_empty = {k: v for k, v in counts.items() if k != "empty"}
        out[h] = max(non_empty.items(), key=lambda x: x[1])[0] if non_empty else "empty"
    return out


def _guess_map(hints: Dict[str, List[str]], headers: List[str], rows: Optional[List[Dict[str, str]]] = None) -> Dict[str, str]:
    """Two-pass mapping: 1) match by header substring, 2) fill gaps by column value type.

    For example if no column header matches "email" but one column's dominant value
    type is "email", we map `email → that header`. Same for host (ip/domain),
    web_url (url), password (password)."""
    out: Dict[str, str] = {}
    used: set = set()
    lower = [(h, (h or "").lower()) for h in headers]

    # Pass 1 — header-name match
    for field, keys in hints.items():
        for orig, low in lower:
            if orig in used:
                continue
            if any(k in low for k in keys):
                out[field] = orig
                used.add(orig)
                break

    # Pass 2 — value-type match.
    # Order matters: required fields (host, email) are filled BEFORE optional ones
    # (web_url, password) so a single matching column goes to the required slot.
    if rows:
        types = column_types(headers, rows)
        # Priority list: (entity_field, accepted_value_types)
        # Earlier entries win the column over later ones.
        priority = [
            ("email",    ("email",)),
            ("host",     ("ip", "domain", "url")),     # URL doubles as host (importer extracts domain)
            ("password", ("password",)),
            ("web_url",  ("url",)),                     # Only after host has had its chance
            ("phone",    ("phone",)),
        ]
        for field, accepted in priority:
            if field not in hints or field in out:
                continue
            for header, t in types.items():
                if header in used: continue
                if t in accepted:
                    out[field] = header
                    used.add(header)
                    break

    return out


def detect_no_header(grid: List[List[str]]) -> bool:
    """Heuristic: if the first row's cells look like real data (≥2 cells are
    email/ip/url/domain/password), it's not a header row."""
    if len(grid) < 2:
        return False
    first = grid[0]
    data_types = {"email", "ip", "url", "domain", "password"}
    data_count = sum(1 for v in first if classify_value(v) in data_types)
    return data_count >= 2


# ── Per-row classification (for the "auto" target) ──────────────────────

def classify_row(row: Dict[str, str], col_types: Dict[str, str]) -> str:
    """Decide what entity a single row should go to: 'servers' | 'mail' | 'skip'.
    Looks at the actual cell values rather than column names.

    Rule:
      • IP / domain / URL  → 'servers' (we can extract host from URL on import)
      • Just email         → 'mail'
      • Neither            → 'skip'
    """
    has_email = False
    has_host  = False
    for v in row.values():
        t = classify_value(v)
        if   t == "email":                      has_email = True
        elif t in ("ip", "domain", "url"):      has_host  = True
    if has_host:  return "servers"
    if has_email: return "mail"
    return "skip"


# ── Host / IP extraction ────────────────────────────────────────────────

IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b(?::\d{1,5})?")
HOST_RE = re.compile(r"\b([a-z0-9][a-z0-9\-]{0,62}\.[a-z0-9\-.]+)\b", re.I)


def _split_host_port(s: str) -> tuple[str, int]:
    s = (s or "").strip()
    if ":" in s:
        h, _, p = s.partition(":")
        try: return h, int(p)
        except Exception: return h, 22
    return s, 22


def _extract_host(value: str) -> Optional[str]:
    if not value: return None
    m = IP_RE.search(value)
    if m: return m.group(0)
    m = HOST_RE.search(value)
    if m: return m.group(0)
    return None


# ── Endpoints ───────────────────────────────────────────────────────────

@router.post("/discover")
async def discover(data: DiscoverIn, user: User = Depends(get_current_user)):
    try:
        return await discover_tabs(data.url)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати таблицю: {e}")


def _analyze(parsed: Dict[str, Any], tab_name: str = "", limit: int = 20) -> Dict[str, Any]:
    """Full per-tab analysis used by all preview endpoints. Returns headers +
    sample rows + column types + per-target auto-mapping + per-row routes +
    aggregated counts + suggested target + capability flags."""
    headers = parsed["headers"]
    rows = parsed["rows"]
    headerless = parsed.get("headerless", False)

    col_types     = column_types(headers, rows)
    guess_servers = _guess_map(SERVER_HINTS, headers, rows)
    guess_mail    = _guess_map(MAIL_HINTS,    headers, rows)
    can_servers   = "host"  in guess_servers
    can_mail      = "email" in guess_mail

    row_routes = [classify_row(r, col_types) for r in rows]
    route_counts = {
        "servers": sum(1 for x in row_routes if x == "servers"),
        "mail":    sum(1 for x in row_routes if x == "mail"),
        "skip":    sum(1 for x in row_routes if x == "skip"),
    }
    can_auto = route_counts["servers"] > 0 or route_counts["mail"] > 0

    if route_counts["servers"] > 0 and route_counts["mail"] > 0:
        suggested = "auto"
    elif route_counts["servers"] > 0 and can_servers:
        suggested = "servers"
    elif route_counts["mail"] > 0 and can_mail:
        suggested = "mail"
    else:
        name_hint = _guess_target_by_tab(tab_name)
        if name_hint == "servers" and can_servers: suggested = "servers"
        elif name_hint == "mail"  and can_mail:    suggested = "mail"
        else: suggested = "notes"

    return {
        "headers": headers,
        "headerless": headerless,
        "rows": rows[:limit],
        "total_rows": parsed["total_rows"],
        "column_types": col_types,
        "guess": {"servers": guess_servers, "mail": guess_mail},
        "route_counts": route_counts,
        "suggested_target": suggested,
        "can_servers": can_servers,
        "can_mail":    can_mail,
        "can_auto":    can_auto,
    }


@router.post("/preview")
async def preview(data: PreviewIn, user: User = Depends(get_current_user)):
    sid = extract_sheet_id(data.url)
    try:
        csv_text = await fetch_tab_csv(sid, data.gid)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось завантажити CSV: {e}")
    parsed = parse_csv(csv_text)
    return _analyze(parsed, tab_name="", limit=data.limit)


async def _run_one_tab(db: AsyncSession, user: User, parsed: Dict[str, Any], target: str,
                       column_map: Dict[str, str], tab_name: str) -> Dict[str, Any]:
    """Shared run-once logic supporting all targets including 'auto'."""
    headers = parsed["headers"]
    rows = parsed["rows"]

    if target == "auto":
        col_types = column_types(headers, rows)
        server_rows = [r for r in rows if classify_row(r, col_types) == "servers"]
        mail_rows   = [r for r in rows if classify_row(r, col_types) == "mail"]
        skipped_rows = len(rows) - len(server_rows) - len(mail_rows)
        created = updated = sk = 0; errors: List[str] = []
        split = []
        if server_rows:
            cmap = _guess_map(SERVER_HINTS, headers, server_rows)
            if "host" in cmap:
                r1 = await _import_servers(db, user, server_rows, cmap)
                split.append({"target": "servers", "rows": len(server_rows), "result": r1})
                created += r1.get("created", 0); updated += r1.get("updated", 0); sk += r1.get("skipped", 0); errors += r1.get("errors", []) or []
            else:
                sk += len(server_rows)
                errors.append("Не знайдено колонку host/ip для серверних рядків")
        if mail_rows:
            cmap = _guess_map(MAIL_HINTS, headers, mail_rows)
            if "email" in cmap:
                r2 = await _import_mail(db, user, mail_rows, cmap)
                split.append({"target": "mail", "rows": len(mail_rows), "result": r2})
                created += r2.get("created", 0); updated += r2.get("updated", 0); sk += r2.get("skipped", 0); errors += r2.get("errors", []) or []
            else:
                sk += len(mail_rows)
                errors.append("Не знайдено колонку email для поштових рядків")
        sk += skipped_rows
        return {"ok": True, "created": created, "updated": updated, "skipped": sk,
                "errors": errors[:5], "split": split}

    if target == "servers":
        return await _import_servers(db, user, rows, column_map)
    if target == "mail":
        return await _import_mail(db, user, rows, column_map)
    if target == "notes":
        return await _import_notes(db, user, headers, rows, tab_name or "Imported sheet")
    raise HTTPException(400, "target має бути servers / mail / notes / auto")


@router.post("/run")
async def run(data: RunIn, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    if data.target not in ("servers", "mail", "notes", "auto"):
        raise HTTPException(400, "target має бути servers / mail / notes / auto")
    sid = extract_sheet_id(data.url)
    try:
        parsed = await fetch_tab(sid, data.gid)
    except Exception as e:
        raise HTTPException(400, f"CSV: {e}")
    return await _run_one_tab(db, user, parsed, data.target, data.column_map, data.tab_name or "")


# ── Batch (multi-tab) Google Sheets import ─────────────────────────────

import asyncio as _asyncio


def _guess_target_by_tab(name: str) -> str:
    lc = (name or "").lower()
    if any(kw in lc for kw in ("vps", "server", "сервер")):       return "servers"
    if any(kw in lc for kw in ("mail", "пошт", "почт", "email")): return "mail"
    return "notes"


@router.post("/batch-discover")
async def batch_discover(data: BatchDiscoverIn, user: User = Depends(get_current_user)):
    """Deep analysis of a Google Sheet:
    - fetches each tab's CSV in parallel
    - detects headerless tabs and synthesizes col1/col2/…
    - classifies each column by VALUE TYPE (email, ip, url, password, …) not just header text
    - auto-maps columns to entity fields using header text + dominant value type
    - per-row routes (servers / mail / skip) for the "auto-split" target
    - returns predicted counts so the UI shows exactly what will be created
    """
    try:
        info = await discover_tabs(data.url)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати таблицю: {e}")
    sid = info["sheet_id"]

    sem = _asyncio.Semaphore(5)

    async def one(tab: dict) -> dict:
        async with sem:
            try:
                parsed = await fetch_tab(sid, tab["gid"])
                headers = parsed["headers"]
                rows = parsed["rows"]

                # Column-level type classification (looks at actual data)
                col_types = column_types(headers, rows)

                # Two-pass auto-mapping (header + value type)
                guess_servers = _guess_map(SERVER_HINTS, headers, rows)
                guess_mail    = _guess_map(MAIL_HINTS,    headers, rows)
                can_servers = "host"  in guess_servers
                can_mail    = "email" in guess_mail

                # Per-row classification → counts for "auto" target
                row_routes = [classify_row(r, col_types) for r in rows]
                route_counts = {
                    "servers": sum(1 for x in row_routes if x == "servers"),
                    "mail":    sum(1 for x in row_routes if x == "mail"),
                    "skip":    sum(1 for x in row_routes if x == "skip"),
                }

                # Suggested target: prefer auto if mixed types, else single
                if route_counts["servers"] > 0 and route_counts["mail"] > 0:
                    suggested = "auto"
                elif route_counts["servers"] > 0 and can_servers:
                    suggested = "servers"
                elif route_counts["mail"] > 0 and can_mail:
                    suggested = "mail"
                else:
                    # Fall back to name-based hint
                    name_hint = _guess_target_by_tab(tab["name"])
                    if name_hint == "servers" and can_servers: suggested = "servers"
                    elif name_hint == "mail"  and can_mail:    suggested = "mail"
                    else: suggested = "notes"

                return {
                    "gid": tab["gid"], "name": tab["name"], "index": tab.get("index", 0),
                    "headers": headers,
                    "headerless": parsed.get("headerless", False),
                    "rows_sample": rows[: data.preview_limit],
                    "total_rows": parsed["total_rows"],
                    "column_types": col_types,
                    "guess": {"servers": guess_servers, "mail": guess_mail},
                    "route_counts": route_counts,
                    "suggested_target": suggested,
                    "can_servers": can_servers,
                    "can_mail": can_mail,
                    "can_auto": route_counts["servers"] > 0 or route_counts["mail"] > 0,
                    "error": None,
                }
            except Exception as e:
                return {
                    "gid": tab["gid"], "name": tab["name"], "index": tab.get("index", 0),
                    "headers": [], "headerless": False, "rows_sample": [], "total_rows": 0,
                    "column_types": {}, "guess": {"servers": {}, "mail": {}},
                    "route_counts": {"servers": 0, "mail": 0, "skip": 0},
                    "suggested_target": "notes", "can_servers": False, "can_mail": False,
                    "can_auto": False,
                    "error": str(e)[:200],
                }

    tabs_data = await _asyncio.gather(*[one(t) for t in info["tabs"]])
    return {
        "sheet_id": sid,
        "title": info.get("title") or "",
        "tabs": tabs_data,
    }


@router.post("/batch-run")
async def batch_run(data: BatchRunIn, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """Import multiple tabs in one call. Returns per-tab result + totals.

    Special target 'auto' = row-by-row split: each row routed to servers/mail/skip
    based on the actual cell values, so a single mixed tab can populate both.
    """
    sid = extract_sheet_id(data.url)
    out_items: List[dict] = []
    totals = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}

    for item in data.items:
        if item.target not in ("servers", "mail", "notes", "auto"):
            out_items.append({"gid": item.gid, "tab_name": item.tab_name, "target": item.target,
                              "ok": False, "error": "Невідома сутність"})
            totals["errors"] += 1
            continue
        try:
            parsed = await fetch_tab(sid, item.gid)
        except Exception as e:
            out_items.append({"gid": item.gid, "tab_name": item.tab_name, "target": item.target,
                              "ok": False, "error": f"Завантаження CSV: {e}"})
            totals["errors"] += 1
            continue
        rows = parsed["rows"]; headers = parsed["headers"]
        try:
            if item.target == "auto":
                col_types = column_types(headers, rows)
                server_rows = [r for r in rows if classify_row(r, col_types) == "servers"]
                mail_rows   = [r for r in rows if classify_row(r, col_types) == "mail"]
                skipped_rows = len(rows) - len(server_rows) - len(mail_rows)
                created = updated = sk = 0; errors: List[str] = []
                split_results = []
                if server_rows:
                    cmap = _guess_map(SERVER_HINTS, headers, server_rows)
                    if "host" in cmap:
                        r1 = await _import_servers(db, user, server_rows, cmap)
                        split_results.append({"target": "servers", "rows": len(server_rows), "result": r1})
                        created += r1.get("created", 0); updated += r1.get("updated", 0); sk += r1.get("skipped", 0); errors += r1.get("errors", []) or []
                    else:
                        sk += len(server_rows)
                        errors.append("Не знайдено колонку host/ip для серверних рядків")
                if mail_rows:
                    cmap = _guess_map(MAIL_HINTS, headers, mail_rows)
                    if "email" in cmap:
                        r2 = await _import_mail(db, user, mail_rows, cmap)
                        split_results.append({"target": "mail", "rows": len(mail_rows), "result": r2})
                        created += r2.get("created", 0); updated += r2.get("updated", 0); sk += r2.get("skipped", 0); errors += r2.get("errors", []) or []
                    else:
                        sk += len(mail_rows)
                        errors.append("Не знайдено колонку email для поштових рядків")
                sk += skipped_rows
                res = {"ok": True, "created": created, "updated": updated, "skipped": sk,
                       "errors": errors[:5], "split": split_results}
            elif item.target == "servers":
                res = await _import_servers(db, user, rows, item.column_map)
            elif item.target == "mail":
                res = await _import_mail(db, user, rows, item.column_map)
            else:
                res = await _import_notes(db, user, headers, rows, item.tab_name or "Imported sheet")
            out_items.append({"gid": item.gid, "tab_name": item.tab_name, "target": item.target,
                              "ok": True, "result": res})
            totals["created"] += res.get("created", 0) or 0
            totals["updated"] += res.get("updated", 0) or 0
            totals["skipped"] += res.get("skipped", 0) or 0
            if res.get("errors"):
                totals["errors"] += len(res["errors"])
        except HTTPException as he:
            out_items.append({"gid": item.gid, "tab_name": item.tab_name, "target": item.target,
                              "ok": False, "error": he.detail})
            totals["errors"] += 1
        except Exception as e:
            out_items.append({"gid": item.gid, "tab_name": item.tab_name, "target": item.target,
                              "ok": False, "error": str(e)[:200]})
            totals["errors"] += 1

    return {"ok": True, "items": out_items, "totals": totals}


# ── Local fortune-sheet workbook parsing ───────────────────────────────

def _parse_fortune_workbook(raw: str, sheet_index: int = 0) -> Dict[str, Any]:
    """Read one sheet of a fortune-sheet workbook into headers+rows."""
    try:
        wb = json.loads(raw or "[]")
    except Exception:
        return {"headers": [], "rows": [], "total_rows": 0, "sheets": []}
    if not isinstance(wb, list) or not wb:
        return {"headers": [], "rows": [], "total_rows": 0, "sheets": []}
    sheets_meta = [{"index": i, "name": s.get("name") or f"Sheet{i+1}"} for i, s in enumerate(wb)]
    if sheet_index < 0 or sheet_index >= len(wb):
        sheet_index = 0
    sheet = wb[sheet_index]
    cd = sheet.get("celldata") or []
    if not cd:
        return {"headers": [], "rows": [], "total_rows": 0, "sheets": sheets_meta}
    max_r = max(c["r"] for c in cd)
    max_c = max(c["c"] for c in cd)
    grid: List[List[str]] = [["" for _ in range(max_c + 1)] for _ in range(max_r + 1)]
    for cell in cd:
        v = cell.get("v") or {}
        val = v.get("v") if isinstance(v, dict) else v
        grid[cell["r"]][cell["c"]] = "" if val is None else str(val)
    # First non-empty row → header candidate
    header_idx = 0
    for i, row in enumerate(grid):
        if any(x.strip() for x in row):
            header_idx = i; break

    candidate = grid[header_idx]
    data_cells = 0
    for c in candidate:
        s = (c or "").strip()
        if not s: continue
        if classify_value(s) in ("email", "ip", "url", "domain", "password"):
            data_cells += 1
    headerless = data_cells >= 2

    if headerless:
        max_cols = max(len(r) for r in grid[header_idx:])
        headers = [f"col{j+1}" for j in range(max_cols)]
        body = grid[header_idx:]
    else:
        headers_raw = candidate
        headers: List[str] = []; seen: Dict[str, int] = {}
        for j, h in enumerate(headers_raw):
            h = (h or "").strip() or f"col{j+1}"
            if h in seen:
                seen[h] += 1; h = f"{h}_{seen[h]}"
            else:
                seen[h] = 1
            headers.append(h)
        body = grid[header_idx + 1:]

    rows: List[Dict[str, str]] = []
    for row in body:
        if not any(x.strip() for x in row):
            continue
        rows.append({headers[i]: (row[i].strip() if i < len(row) else "") for i in range(len(headers))})
    return {"headers": headers, "rows": rows, "total_rows": len(rows), "headerless": headerless, "sheets": sheets_meta}


async def _load_local_sheet(db: AsyncSession, sheet_id: int, user: User) -> Spreadsheet:
    s = await db.get(Spreadsheet, sheet_id)
    if not s or s.owner_user_id != user.id:
        raise HTTPException(404, "Таблиця не знайдена")
    return s


@router.post("/local/preview")
async def local_preview(data: LocalPreviewIn, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    s = await _load_local_sheet(db, data.sheet_id, user)
    parsed = _parse_fortune_workbook(s.data or "[]", data.sheet_index)
    analysis = _analyze(parsed, tab_name=s.name)
    analysis["sheets"]     = parsed.get("sheets", [])
    analysis["sheet_name"] = s.name
    return analysis


@router.post("/local/run")
async def local_run(data: LocalRunIn, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    if data.target not in ("servers", "mail", "notes", "auto"):
        raise HTTPException(400, "target має бути servers / mail / notes / auto")
    s = await _load_local_sheet(db, data.sheet_id, user)
    parsed = _parse_fortune_workbook(s.data or "[]", data.sheet_index)
    return await _run_one_tab(db, user, parsed, data.target, data.column_map, data.tab_name or s.name)


# ── File-upload import (xlsx/csv, one-shot, nothing stored) ────────────

def _parse_xlsx_first_sheet(buf: bytes) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl не встановлено на бекенді")
    wb = load_workbook(io.BytesIO(buf), data_only=True, read_only=True)
    sheets_meta = [{"index": i, "name": n} for i, n in enumerate(wb.sheetnames)]
    return {"workbook": wb, "sheets": sheets_meta}


def _xlsx_sheet_to_rows(wb, sheet_index: int) -> Dict[str, Any]:
    names = wb.sheetnames
    if sheet_index < 0 or sheet_index >= len(names):
        sheet_index = 0
    ws = wb[names[sheet_index]]
    grid: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append(["" if v is None else str(v).strip() for v in row])
    if not grid:
        return {"headers": [], "rows": [], "total_rows": 0, "headerless": False}
    header_idx = 0
    for i, row in enumerate(grid):
        if any(x for x in row):
            header_idx = i; break

    # Headerless detection — same rules as parse_csv
    candidate = grid[header_idx]
    data_cells = 0
    for c in candidate:
        s = (c or "").strip()
        if not s: continue
        t = classify_value(s)
        if t in ("email", "ip", "url", "domain", "password"):
            data_cells += 1
    headerless = data_cells >= 2

    if headerless:
        max_cols = max((len(r) for r in grid[header_idx:]), default=0)
        headers = [f"col{j+1}" for j in range(max_cols)]
        body = grid[header_idx:]
    else:
        headers_raw = candidate
        headers: List[str] = []; seen: Dict[str, int] = {}
        for j, h in enumerate(headers_raw):
            h = (h or "").strip() or f"col{j+1}"
            if h in seen: seen[h] += 1; h = f"{h}_{seen[h]}"
            else: seen[h] = 1
            headers.append(h)
        body = grid[header_idx + 1:]

    rows: List[Dict[str, str]] = []
    for row in body:
        if not any(x for x in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    return {"headers": headers, "rows": rows, "total_rows": len(rows), "headerless": headerless}


@router.post("/file/preview")
async def file_preview(file: UploadFile = File(...), sheet_index: int = 0, user: User = Depends(get_current_user)):
    raw = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".csv") or file.content_type == "text/csv":
        parsed = parse_csv(raw.decode("utf-8", errors="replace"))
        sheets = [{"index": 0, "name": file.filename or "CSV"}]
    else:
        info = _parse_xlsx_first_sheet(raw)
        parsed = _xlsx_sheet_to_rows(info["workbook"], sheet_index)
        sheets = info["sheets"]
    analysis = _analyze(parsed, tab_name=file.filename or "")
    analysis["sheets"]     = sheets
    analysis["sheet_name"] = file.filename
    return analysis


@router.post("/file/run")
async def file_run(
    file: UploadFile = File(...),
    target: str = "notes",
    column_map: str = "{}",
    sheet_index: int = 0,
    tab_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if target not in ("servers", "mail", "notes", "auto"):
        raise HTTPException(400, "target має бути servers / mail / notes / auto")
    try:
        cmap = json.loads(column_map or "{}")
    except Exception:
        raise HTTPException(400, "column_map має бути JSON")
    raw = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".csv") or file.content_type == "text/csv":
        parsed = parse_csv(raw.decode("utf-8", errors="replace"))
    else:
        info = _parse_xlsx_first_sheet(raw)
        parsed = _xlsx_sheet_to_rows(info["workbook"], sheet_index)
    return await _run_one_tab(db, user, parsed, target, cmap, tab_name or file.filename or "Imported file")


# ── Importers ────────────────────────────────────────────────────────────

def _get(row: Dict[str, str], cmap: Dict[str, str], field: str) -> str:
    header = cmap.get(field)
    if not header:
        return ""
    return (row.get(header) or "").strip()


async def _import_servers(db: AsyncSession, user: User, rows: List[Dict[str, str]], cmap: Dict[str, str]):
    if "host" not in cmap:
        raise HTTPException(400, "Потрібно вказати колонку для host/ip")
    existing = (await db.execute(
        select(RemoteServer).where(RemoteServer.owner_user_id == user.id)
    )).scalars().all()
    by_host = {(s.host or "").lower(): s for s in existing}

    created = updated = skipped = 0
    errors: List[str] = []

    for row in rows:
        try:
            raw_host = _get(row, cmap, "host")
            host_clean = _extract_host(raw_host)
            if not host_clean:
                skipped += 1
                continue
            host, port = _split_host_port(host_clean if ":" in raw_host else raw_host) \
                if IP_RE.search(raw_host or "") else (host_clean, 22)
            host = host.lower()
            label    = _get(row, cmap, "label")    or host
            username = _get(row, cmap, "username") or "root"
            password = _get(row, cmap, "password")
            web_url  = _get(row, cmap, "web_url")
            notes    = _get(row, cmap, "notes")
            tags     = _get(row, cmap, "tags")

            obj = by_host.get(host)
            if obj is None:
                obj = RemoteServer(
                    owner_user_id=user.id,
                    label=label, host=host, port=port,
                    username=username, auth_kind="password",
                    password_enc=encrypt_secret(password) if password else None,
                    web_url=web_url or None,
                    tags=tags or None,
                    notes=notes or None,
                )
                db.add(obj)
                by_host[host] = obj
                created += 1
            else:
                changed = False
                if label and obj.label != label: obj.label = label; changed = True
                if username and obj.username != username: obj.username = username; changed = True
                if password:
                    obj.password_enc = encrypt_secret(password); changed = True
                if web_url and obj.web_url != web_url: obj.web_url = web_url; changed = True
                if notes and obj.notes != notes: obj.notes = notes; changed = True
                if tags and obj.tags != tags: obj.tags = tags; changed = True
                if changed: updated += 1
                else: skipped += 1
        except Exception as e:
            errors.append(str(e)[:200])
            skipped += 1

    await db.flush()
    return {"ok": True, "target": "servers", "created": created, "updated": updated,
            "skipped": skipped, "errors": errors[:5]}


async def _import_mail(db: AsyncSession, user: User, rows: List[Dict[str, str]], cmap: Dict[str, str]):
    if "email" not in cmap:
        raise HTTPException(400, "Потрібно вказати колонку для email")
    existing = (await db.execute(
        select(MailAccount).where(MailAccount.owner_user_id == user.id)
    )).scalars().all()
    by_email = {(m.email or "").lower(): m for m in existing}

    created = updated = skipped = 0
    errors: List[str] = []

    for row in rows:
        try:
            email = _get(row, cmap, "email").lower()
            if not email or "@" not in email:
                skipped += 1
                continue
            password = _get(row, cmap, "password")
            label    = _get(row, cmap, "label")
            tags     = _get(row, cmap, "tags")
            notes    = _get(row, cmap, "notes")

            domain = email.split("@", 1)[1]
            # Reasonable IMAP defaults — user can edit per account later.
            imap_host = f"imap.{domain}"

            obj = by_email.get(email)
            if obj is None:
                if not password:
                    skipped += 1
                    continue
                obj = MailAccount(
                    owner_user_id=user.id,
                    label=label or email, email=email,
                    imap_host=imap_host, imap_port=993, imap_ssl=True,
                    username=email,
                    password_enc=encrypt_secret(password),
                    tags=tags or None,
                    notes=notes or None,
                )
                db.add(obj)
                by_email[email] = obj
                created += 1
            else:
                changed = False
                if password:
                    obj.password_enc = encrypt_secret(password); changed = True
                if label and obj.label != label: obj.label = label; changed = True
                if tags and obj.tags != tags: obj.tags = tags; changed = True
                if notes and obj.notes != notes: obj.notes = notes; changed = True
                if changed: updated += 1
                else: skipped += 1
        except Exception as e:
            errors.append(str(e)[:200])
            skipped += 1

    await db.flush()
    return {"ok": True, "target": "mail", "created": created, "updated": updated,
            "skipped": skipped, "errors": errors[:5]}


async def _import_notes(db: AsyncSession, user: User, headers: List[str], rows: List[Dict[str, str]], title: str):
    """Dump tab as a markdown table inside a single Note."""
    if not rows:
        return {"ok": True, "target": "notes", "created": 0, "skipped": 0, "rows": 0}
    md_lines = ["| " + " | ".join(headers) + " |",
                "| " + " | ".join(["---"] * len(headers)) + " |"]
    for r in rows:
        md_lines.append("| " + " | ".join((r.get(h, "") or "").replace("|", "\\|") for h in headers) + " |")
    body = "\n".join(md_lines)

    n = Note(
        owner_user_id=user.id,
        title=title[:200],
        body=body,
    )
    db.add(n)
    await db.flush()
    return {"ok": True, "target": "notes", "created": 1, "note_id": n.id, "rows": len(rows)}
